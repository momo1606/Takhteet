from __future__ import unicode_literals
import os
import pandas
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import math



month=["محرم الحرام","صفر المظفر","ربيع الأول","ربيع الأخر","جمادي الأول","جمادي الأخر","رجب الأصب","شعبان الكريم","رمضان المعظم","شوال المكرم","ذي القعدة الحرام","ذي الحجة الحرام"]

sup=[None,"الفاتحة","البقرة","آل عمران","النساء","المائدة","الأنعام","الأعراف","الأنفال","التوبة","يونس","هود","يوسف","الرعد","إبراهيم","الحجر","النحل","الإسراء","الكهف",
       "مريم","طه","الأنبياء","الحج","المؤمنون","النور","الفرقان","الشعراء","النمل","القصص","العنكبوت","الروم","لقمان","السجدة","الأحزاب","سبأ","فاطر","يس","الصافات",
       "ص","الزمر","غافر","فصلت","الشورى","الزخرف","الدخان","الجاثية","الأحقاف","محمد","الفتح","الحجرات","ق","الذاريات","الطور","النجم","القمر","الرحمن","الواقعة",
       "الحديد","المجادلة","الحشر","الممتحنة","الصف","الجمعة","المنافقون","التغابن","الطلاق","التحريم","الملك","القلم","الحاقة","المعارج","نوح","الجن","المزمل",
       "المدثر","القيامة","الإنسان","المرسلات","النبأ","النازعات","عبس","التكوير","الإنفطار","المطففين","الانشقاق","البروج","الطارق","الأعلى","الغاشية","الفجر","البلد",
       "الشمس","الليل","الضحى","الشرح","التين","العلق","القدر","البينة","الزلزلة","العاديات","القارعة","التكاثر","العصر","الهمزة","الفيل","قريش","الماعون","الكوثر",
       "الكافرون","النصر","المسد","الإخلاص","الفلق","الناس"]


def to_arab(q):
    def enToArNumb(number): 
        dic = { 
            0:'۰', 
            1:'١', 
            2:'٢', 
            3:'۳', 
            4:'٤', 
            5:'۵', 
            6:'٦',
            7:'۷', 
            8:'۸', 
            9:'۹',  
        }
        return dic.get(number)
    fin=""
    while(q>0):
        a=q%10
        q=q//10
        l=enToArNumb(a)
        fin=fin+l
    return (fin[::-1])
def to_excel2(q_list,days,form,name,n,month,yr):
    l1,l2,l3,l4,l5,l6,l7,l8,l9,l10,l11,l12,l13,l14,l15,l16,l17,l18=[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
    if(form==2):
        for i in range(0,len(q_list)):
            j=i+1
            if((i%6)==0):
                l1.append(" "+to_arab(j)+" ")
                l1.append(to_arab(q_list[i][1])+" : "+to_arab(q_list[i][0]))
                l1.append(None)
                l1.append("الإمضاء")
            elif((i%6)==1):
                l2.append(" "+to_arab(j)+" ")
                l2.append(to_arab(q_list[i][1])+" : "+to_arab(q_list[i][0]))
                l2.append(None)
                l2.append("الإمضاء")
            elif((i%6)==2):
                l3.append(" "+to_arab(j)+" ")
                l3.append(to_arab(q_list[i][1])+" : "+to_arab(q_list[i][0]))
                l3.append(None)
                l3.append("الإمضاء")
            elif((i%6)==3):
                l4.append(" "+to_arab(j)+" ")
                l4.append(to_arab(q_list[i][1])+" : "+to_arab(q_list[i][0]))
                l4.append(None)
                l4.append("الإمضاء")
            elif((i%6)==4):
                l5.append(" "+to_arab(j)+" ")
                l5.append(to_arab(q_list[i][1])+" : "+to_arab(q_list[i][0]))
                l5.append(None)
                l5.append("الإمضاء")
            elif((i%6)==5):
                l6.append(" "+to_arab(j)+" ")
                l6.append(to_arab(q_list[i][1])+" : "+to_arab(q_list[i][0]))
                l6.append(None)
                l6.append("الإمضاء")
            if(i==len(q_list)-1):
                v=[l1]+[l2]+[l3]+[l4]+[l5]+[l6]
                for i in range(1,len(v)):
                    if(len(v[i])<len(v[0])):
                        v[i].append(None)
                        v[i].append(None)
                        v[i].append(None)
                        v[i].append(None)
    elif(form==1):
        for i in range(0,len(q_list)):
            j=i+1
            if(i==0):
                l7.append(None)
                l8.append(None)
                l9.append(None)
                l10.append(None)
                l11.append(None)
            l1.append(None)
            l5.append(None)
            l6.append(None)
            l2.append(to_arab(q_list[i][1])+" : "+to_arab(q_list[i][0]))
            l3.append(" "+to_arab(j)+" ")
            if(i==len(q_list)-1):
                v=[l1]+[l2]+[l3]+[l4]+[l5]+[l6]
    for i in range(0,len(v)):
        v[i].append(" ")
        v[i].append(" ")
    if(form==1):
        df1 = pandas.DataFrame({"الإسم"+" : "+name:l7, month+' '+yr:l8, "أيام العمل"+" : "+to_arab(len(q_list)): l9, "سطور يوميا"+" : "+to_arab(n):l10, " - تفاصيل":l11})
        df = pandas.DataFrame({"الإمضاء":v[4], "حصول":v[5], "سطر": v[0], "التخطيط": v[1], "تاريخ":v[2]})
    elif(form==2):
        df1 = pandas.DataFrame({name:l12, "الإسم"+" : ":l7, month+' '+yr:l8, "أيام العمل"+" : "+to_arab(len(q_list)): l9, "سطور يوميا"+" : "+to_arab(n):l10, " - تفاصيل":l11})
        df2 = pandas.DataFrame({to_arab(n*len(q_list)):l18, "عدد السطور"+" : ":l17, to_arab(q_list[-1][1])+" - "+to_arab(q_list[-1][0]):l16, " إلى "+" : ": l15, to_arab(q_list[0][1])+" - "+to_arab(q_list[0][0]):l14, " من"+" : ":l13})
        df = pandas.DataFrame({"ra":v[5], "re":v[4], "ri": v[3], "ro": v[2], "ru":v[1], "r":v[0]})
    book = load_workbook('takhteet.xlsx')
    writer = pandas.ExcelWriter('takhteet.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        if(form==1):
            df1.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
        elif(form==2):
            df1.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df2.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)

    writer.save()
def to_excel1(q_list,days,form,name,n,month,yr):
    l1,l2,l3,l4,l5,l6,l7,l8,l9,l10,l11,l12,l13,l14,l15,l16,l17,l18=[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
    if(form==2):
        for i in range(0,len(q_list)):
            j=i+1
            if((i%6)==0):
                l1.append(" "+to_arab(j)+" ")
                l1.append(sup[q_list[i][0]]+" : "+to_arab(q_list[i][1]))
                l1.append(None)
                l1.append("الإمضاء")
            elif((i%6)==1):
                l2.append(" "+to_arab(j)+" ")
                l2.append(sup[q_list[i][0]]+" : "+to_arab(q_list[i][1]))
                l2.append(None)
                l2.append("الإمضاء")
            elif((i%6)==2):
                l3.append(" "+to_arab(j)+" ")
                l3.append(sup[q_list[i][0]]+" : "+to_arab(q_list[i][1]))
                l3.append(None)
                l3.append("الإمضاء")
            elif((i%6)==3):
                l4.append(" "+to_arab(j)+" ")
                l4.append(sup[q_list[i][0]]+" : "+to_arab(q_list[i][1]))
                l4.append(None)
                l4.append("الإمضاء")
            elif((i%6)==4):
                l5.append(" "+to_arab(j)+" ")
                l5.append(sup[q_list[i][0]]+" : "+to_arab(q_list[i][1]))
                l5.append(None)
                l5.append("الإمضاء")
            elif((i%6)==5):
                l6.append(" "+to_arab(j)+" ")
                l6.append(sup[q_list[i][0]]+" : "+to_arab(q_list[i][1]))
                l6.append(None)
                l6.append("الإمضاء")
            if(i==len(q_list)-1):
                v=[l1]+[l2]+[l3]+[l4]+[l5]+[l6]
                for i in range(1,len(v)):
                    if(len(v[i])<len(v[0])):
                        v[i].append(None)
                        v[i].append(None)
                        v[i].append(None)
                        v[i].append(None)
    elif(form==1):
        for i in range(0,len(q_list)):
            j=i+1
            if(i==0):
                l7.append(None)
                l8.append(None)
                l9.append(None)
                l10.append(None)
                l11.append(None)
            l1.append(None)
            l5.append(None)
            l6.append(None)
            l2.append(sup[q_list[i][0]]+" : "+to_arab(q_list[i][1]))
            l3.append(" "+to_arab(j)+" ")
            if(i==len(q_list)-1):
                v=[l1]+[l2]+[l3]+[l4]+[l5]+[l6]
    for i in range(0,len(v)):
        v[i].append(" ")
        v[i].append(" ")
    y=n*len(q_list)
    if(n-math.floor(n)==0):
        pages_per_day=to_arab(int(n))+'٫'+'۰۰'
    else:
        pages_per_day=to_arab(int(n))+'٫'+to_arab(int(round((n-math.floor(n)),3)*100))
    if(y-math.floor(y)==0):
        total_pages=to_arab(int(y))+'٫'+'۰۰'
    else:
        total_pages=to_arab(int(y))+'٫'+to_arab(int(round((y-math.floor(y)),3)*100))
    if(form==1):
        df1 = pandas.DataFrame({"الإسم"+" : "+name:l7, month+' '+yr:l8, "أيام العمل"+" : "+to_arab(len(q_list)): l9, "صفحات يوميا"+" : "+pages_per_day:l10, " - تفاصيل":l11})
        df = pandas.DataFrame({"الإمضاء":v[4], "حصول":v[5], "صفحة": v[0], "التخطيط": v[1], "تاريخ":v[2]})
    elif(form==2):
        df1 = pandas.DataFrame({name:l12, "الإسم"+" : ":l7, month+' '+yr:l8, "أيام العمل"+" : "+to_arab(len(q_list)): l9, "صفحات يوميا"+" : "+pages_per_day:l10, " - تفاصيل":l11})
        df2 = pandas.DataFrame({total_pages:l18, "عدد الصفحات"+" : ":l17, sup[q_list[-1][0]]+" - "+to_arab(q_list[-1][1]):l16, " إلى "+" : ": l15, sup[q_list[0][0]]+" - "+to_arab(q_list[0][1]):l14, " من"+" : ":l13})
        df = pandas.DataFrame({"ra":v[5], "re":v[4], "ri": v[3], "ro": v[2], "ru":v[1], "r":v[0]})
    book = load_workbook('takhteet.xlsx')
    '''font = Font(name='DecoType Naskh Variants',
		size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                 strike=False,
               color='FF000000')
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))'''
    writer = pandas.ExcelWriter('takhteet.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    

    for sheetname in writer.sheets:
        if(form==1):
            df1.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
        elif(form==2):
            df1.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df2.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)

    writer.save()


                      
'''if __name__ == '__main__':
    c=input("Enter name: ")
    to_excel1([[2,7],[3,45]],2,2,c,2.50,"may","2020")
    to_excel2([[2,7],[3,45]],2,1,c,2,"may","2020")'''

    
