from __future__ import unicode_literals
import numpy as np
import os
import pandas
import time
from openpyxl import Workbook
from openpyxl import load_workbook
a = np.array([[7], [5,16,24,29,37,48,57,61,69,76,83,88,93,101,105,112,119,126,134,141,145,153,163,169,176,181,186,190,196,202,210,215,219,224,230,233,237,245,248,252,256,259,264,269,274,281,282,286],
             [9,15,22,29,37,45,52,61,70,77,83,91,100,108,115,121,132,140,148,153,157,165,173,180,186,194,200],[6,11,14,19,23,26,33,37,44,51,59,65,74,79,86,91,94,101,105,113,121,127,134,140,147,154,162,170,175],
              [2,55,9,13,17,23,31,36,41,45,50,57,64,70,76,82,89,95,103,108,113,120],[8,18,27,35,44,52,59,68,73,81,90,94,101,110,118,124,131,137,142,146,151,157,165],
              [11,22,30,37,43,51,57,67,73,81,87,95,104,120,130,137,143,149,155,159,163,170,178,187,195,206],[8,16,25,33,40,45,52,61,69,75],[6,13,20,26,31,36,40,47,54,61,68,72,79,86,93,99,106,111,117,122,129],
              [6,14,20,25,33,42,53,61,70,78,88,97,106],[5,12,19,28,37,45,53,62,71,81,88,97,108,117],[4,14,22,30,37,43,52,63,69,78,86,95,103,111,],[5,13,18,28,34,42],[5,10,18,24,33,42,52],
              [15,31,51,70,90],[6,14,26,34,42,54,64,72,79,87,93,102,110,118,128],[7,17,27,38,49,58,66,75,86,96,104],[4,15,20,27,34,45,53,61,74,83,97,110],[11,25,38,51,64,76,95],[12,37,51,64,76,87,98,113,125,135],
              [10,24,35,44,57,72,81,90,101,112],[5,15,23,30,38,46,55,64,72,78],[17,27,42,59,74,89,104,118],[10,20,27,31,36,43,53,58,61],[2,11,20,32,43,55,67,77],[19,39,60,83,111,136,159,183,206,227],
              [13,22,35,44,55,63,76,88],[5,13,21,28,35,43,50,59,70,77,84],[6,14,23,30,38,45,52,63],[5,15,24,32,41,50,60],[11,19,28,34],[11,20,30],[6,15,22,30,35,43,50,54,62,73],[7,14,22,31,39,48],
              [3,11,18,30,38,44],[12,27,40,54,70,83],[24,51,76,102,126,153,182],[16,26,42,61,83],[5,10,21,31,40,47,56,67,74],[7,16,25,33,40,49,58,66,77,85],[11,20,29,38,46,54],[10,15,22,31,44,51],
              [10,22,33,47,60,73,89],[18,39,59],[13,22,32],[5,14,20,28,35],[11,19,29,38],[9,15,23,28],[4,11,18],[15,35],[6,30,51],[14,31,49],[26,44],[6,27,49],[16,40,67],[16,50,76],[3,11,18,24,29],
              [6,11,21],[3,9,16,24],[5,11],[5,14],[8],[4,11],[9,18],[5,12],[7,12],[12,26],[15,42],[8,34],[10,39],[10,28],[13,28],[19],[17,47],[19],[5,25],[19,50],[30],[15,46],[42],[29],[6,34],[25],[22],
              [15],[26],[23],[20],[14],[8],[19],[7],[9],[8],[5],[3],[5],[6]])
#print(len(a))
#print(len(a[len(a)-1]))
#print(a[1].index(24))
s=["الفاتحة","البقرة","آل عمران","النساء","المائدة","الأنعام","الأعراف","الأنفال","التوبة","يونس","هود","يوسف","الرعد","إبراهيم","الحجر","النحل","الإسراء","الكهف",
       "مريم","طه","الأنبياء","الحج","المؤمنون","النور","الفرقان","الشعراء","النمل","القصص","العنكبوت","الروم","لقمان","السجدة","الأحزاب","سبأ","فاطر","يس","الصافات",
       "ص","الزمر","غافر","فصلت","الشورى","الزخرف","الدخان","الجاثية","الأحقاف","محمد","الفتح","الحجرات","ق","الذاريات","الطور","النجم","القمر","الرحمن","الواقعة",
       "الحديد","المجادلة","الحشر","الممتحنة","الصف","الجمعة","المنافقون","التغابن","الطلاق","التحريم","الملك","القلم","الحاقة","المعارج","نوح","الجن","المزمل",
       "المدثر","القيامة","الإنسان","المرسلات","النبأ","النازعات","عبس","التكوير","المطففين","الانشقاق","البروج","الأعلى","الغاشية","الفجر","البلد",
       "الليل","الشرح","العلق","البينة","العاديات","التكاثر","الفيل","الكوثر","المسد","الناس"]
month=["محرم الحرام","صفر المظفر","ربيع الأول","ربيع الأخر","جمادي الأول","جمادي الأخر","رجب الأصب","شعبان الكريم","رمضان المعظم","شوال المكرم","ذي القعدة الحرام","ذي الحجة الحرام"]

def surat(sur):
    
    return s[sur]
def ayat(q):
    def enToArNumb(number): 
        dic = { 
            0:'۰', 
            1:'١', 
            2:'٢', 
            3:'۳', 
            4:'٤', 
            5:'۵', 
            6:'٦',
            7:'۷', 
            8:'۸', 
            9:'۹',  
        }
        return dic.get(number)
    fin=""
    while(q>0):
        a=q%10
        q=q//10
        l=enToArNumb(a)
        fin=fin+l
    return (fin[::-1])
def takhteet(sur,aya):
    
    g=ayat(a[sur][aya])
    h=surat(sur)
    print(g,"\t\t",h,"\n")
    return g,h
def generate(form,sur,aya,no,n,name,par,tno,mont,yr):#sur and aya are indices of current surat and ayat 
    l1=[]
    l2=[]
    l3=[]
    l4=[]
    l5=[]
    l6=[]
    l7=[]
    l8=[]
    l9=[]
    l10=[]
    l11=[]
    l12=[]
    l13=[]
    l14=[]
    l15=[]
    l16=[]
    l17=[]
    l18=[]
    e=1
    for i in range(0,no):  #no is no of days and n is pages per day
        
        for j in range(n):
            c=len(a[sur])-1
            if(a[sur][aya]==a[sur][c]):
                sur+=1
                aya=0
            else:
                aya+=1
        g,h=takhteet(sur,aya)
        if(form==2):
            if(par=='n'):
                if(i==0):
                    f1=(h+" - "+g)
                if(i==(no-1)):
                    f2=(h+" - "+g)
                if((i%6)==0):
                    l1.append(" "+ayat(i+1)+" ")
                    l1.append(h+" : "+g)
                    l1.append(None)
                    l1.append("الإمضاء")
                elif((i%6)==1):
                    l2.append(" "+ayat(i+1)+" ")
                    l2.append(h+" : "+g)
                    l2.append(None)
                    l2.append("الإمضاء")
                elif((i%6)==2):
                    l3.append(" "+ayat(i+1)+" ")
                    l3.append(h+" : "+g)
                    l3.append(None)
                    l3.append("الإمضاء")
                elif((i%6)==3):
                    l4.append(" "+ayat(i+1)+" ")
                    l4.append(h+" : "+g)
                    l4.append(None)
                    l4.append("الإمضاء")
                elif((i%6)==4):
                    l5.append(" "+ayat(i+1)+" ")
                    l5.append(h+" : "+g)
                    l5.append(None)
                    l5.append("الإمضاء")
                elif((i%6)==5):
                    l6.append(" "+ayat(i+1)+" ")
                    l6.append(h+" : "+g)
                    l6.append(None)
                    l6.append("الإمضاء")
            elif(par=='y'):
                if(i==0):
                    f1=(h+" - "+g)
                if(i==(no-1)):
                    f2=(h+" - "+g)
                if((i%3)==0):
                    l1.append(" "+ayat(e)+" ")
                    l1.append(None)
                    l1.append(None)
                    l1.append("الإمضاء")
                    l2.append(" "+ayat(e+1)+" ")
                    l2.append(h+" : "+g)
                    l2.append(None)
                    l2.append("الإمضاء")
                elif((i%3)==1):
                    l3.append(" "+ayat(e)+" ")
                    l3.append(None)
                    l3.append(None)
                    l3.append("الإمضاء")
                    l4.append(" "+ayat(e+1)+" ")
                    l4.append(h+" : "+g)
                    l4.append(None)
                    l4.append("الإمضاء")
                elif((i%3)==2):
                    l5.append(" "+ayat(e)+" ")
                    l5.append(None)
                    l5.append(None)
                    l5.append("الإمضاء")
                    l6.append(" "+ayat(e+1)+" ")
                    l6.append(h+" : "+g)
                    l6.append(None)
                    l6.append("الإمضاء")
                    
        elif(form==1):
            if(par=='n'):
                if(i==0):
                    l7.append(None)
                    l8.append(None)
                    l9.append(None)
                    l10.append(None)
                    l11.append(None)
                l1.append(None)
                l5.append(None)
                l6.append(None)
                l2.append(h+" : "+g)
                l3.append(" "+ayat(i+1)+" ")
                if(i==0):
                    l4.append("الإسم"+" : "+name)
                elif(i==1):
                    l4.append("ربيع الأخر ١٤٤٠")
                elif(i==2):
                    l4.append("أيام العمل"+" : "+ayat(20))
                elif(i==3):
                    l4.append("صفحات يوميا"+" : "+ayat(n))
                else:
                    l4.append(None)
            elif(par=='y'):
                if(i==0):
                    l7.append(None)
                    l8.append(None)
                    l9.append(None)
                    l10.append(None)
                    l11.append(None)
                l1.append(None)
                l5.append(None)
                l6.append(None)
                l1.append(None)
                l5.append(None)
                l6.append(None)
                l2.append(None)
                l2.append(h+" : "+g)
                l3.append(" "+ayat(e)+" ")
                l3.append(" "+ayat(e+1)+" ")
                if(i==0):
                    l4.append("الإسم"+" : "+name)
                elif(i==1):
                    l4.append("ربيع الأخر ١٤٤٠")
                elif(i==2):
                    l4.append("أيام العمل"+" : "+ayat(20))
                elif(i==3):
                    l4.append("صفحات يوميا"+" : "+ayat(n))
                else:
                    l4.append(None)
        e=e+2
    if((form==2) and (par=='n')):
        if(len(l2)<len(l1)):
            l2.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l2.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l2.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l2.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
        elif(len(l3)<len(l1)):
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
        elif(len(l4)<len(l1)):
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
        elif(len(l5)<len(l1)):
            l5.append(None)
            l6.append(None)
            l5.append(None)
            l6.append(None)
            l5.append(None)
            l6.append(None)
            l5.append(None)
            l6.append(None)
        elif(len(l6)<len(l1)):
            l6.append(None)
            l6.append(None)
            l6.append(None)
            l6.append(None)
    elif((form==2) and (par=='y')):
        if(len(l3)<len(l1)):
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
            l3.append(None)
            l4.append(None)
            l5.append(None)
            l6.append(None)
        elif(len(l5)<len(l1)):
            l5.append(None)
            l6.append(None)
            l5.append(None)
            l6.append(None)
            l5.append(None)
            l6.append(None)
            l5.append(None)
            l6.append(None)
    l1.append(" ")
    l2.append(" ")
    l3.append(" ")
    l4.append(" ")
    l5.append(" ")
    l6.append(" ")
    l1.append(" ")
    l2.append(" ")
    l3.append(" ")
    l4.append(" ")
    l5.append(" ")
    l6.append(" ")
    if(form==1):
        df1 = pandas.DataFrame({"الإسم"+" : "+name:l7, mont+yr:l8, "أيام العمل"+" : "+ayat(tno): l9, "صفحات يوميا"+" : "+ayat(n):l10, " - تفاصيل":l11})
        df = pandas.DataFrame({"الإمضاء":l5, "حصول":l6, "صفحة": l1, "التخطيط": l2, "تاريخ":l3})
    elif(form==2):
        df1 = pandas.DataFrame({name:l12, "الإسم"+" : ":l7, mont+yr:l8, "أيام العمل"+" : "+ayat(tno): l9, "صفحات يوميا"+" : "+ayat(n):l10, " - تفاصيل":l11})
        df2 = pandas.DataFrame({ayat(n*no):l18, "عدد الصفحات"+" : ":l17, f2:l16, " إلى "+" : ": l15, f1:l14, " من"+" : ":l13})
        df = pandas.DataFrame({"ra":l6, "re":l5, "ri": l4, "ro": l3, "ru":l2, "r":l1})
    book = load_workbook('takhteet.xlsx')
    writer = pandas.ExcelWriter('takhteet.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        if(form==1):
            df1.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
        elif(form==2):
            df1.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df2.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= True)
            df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)

    writer.save()
    '''writer = pandas.ExcelWriter('test1.xlsx')
    df.to_excel(writer,'Sheet1')
    #df2.to_excel(writer,'Sheet2')
    writer.save()
    #df.to_excel('test1.xlsx', sheet_name='sheet1', index=False)'''
            
def search(arr,x):
    for i in range(len(arr)): 
          
        if(arr[i] == x): 
            return i 
        elif((i==0) and (arr[0]>x)):
            return i
        elif((i==len(arr)-1) and (arr[i]<x)):
            return 53
        elif((arr[i]<x) and (arr[i+1]>x)):
            return i+1
  
    return -1

#print(a[1])
#n=search(a[1],16)
#print(n)
#generate(94,0,1,2)
sup=["الفاتحة","البقرة","آل عمران","النساء","المائدة","الأنعام","الأعراف","الأنفال","التوبة","يونس","هود","يوسف","الرعد","إبراهيم","الحجر","النحل","الإسراء","الكهف",
       "مريم","طه","الأنبياء","الحج","المؤمنون","النور","الفرقان","الشعراء","النمل","القصص","العنكبوت","الروم","لقمان","السجدة","الأحزاب","سبأ","فاطر","يس","الصافات",
       "ص","الزمر","غافر","فصلت","الشورى","الزخرف","الدخان","الجاثية","الأحقاف","محمد","الفتح","الحجرات","ق","الذاريات","الطور","النجم","القمر","الرحمن","الواقعة",
       "الحديد","المجادلة","الحشر","الممتحنة","الصف","الجمعة","المنافقون","التغابن","الطلاق","التحريم","الملك","القلم","الحاقة","المعارج","نوح","الجن","المزمل",
       "المدثر","القيامة","الإنسان","المرسلات","النبأ","النازعات","عبس","التكوير","الإنفطار","المطففين","الانشقاق","البروج","الطارق","الأعلى","الغاشية","الفجر","البلد",
       "الشمس","الليل","الضحى","الشرح","التين","العلق","القدر","البينة","الزلزلة","العاديات","القارعة","التكاثر","العصر","الهمزة","الفيل","قريش","الماعون","الكوثر",
       "الكافرون","النصر","المسد","الإخلاص","الفلق","الناس"]

print("*********************** WELCOME ***************************")
print("USER GUIDELINES")
print("\n1. The user has to feed-in the following details:\n\ta)Number of students\n\tb) Total number of days\n\t and then for each student,\n\tc) Student's name\n\td) Student's current surat number and ayat number\n\te) Number of safahs/day assigned to the student.")
print("\n2. The minimum unit of this generator is one page,\n so if a student has a takhteet of 1.5 pages(progression of 0.5),\n total no.of days will be half of actual value\n and pages/day will be 3 so each output will be a target for 2 days,\n similarly for .75 safah no.of days will be a fourth of actual value,\n pages/day will be 3 hence each output is target of 4 days,\n Special provision has been made for progressions of 0.5 of pages/day.")  
print("\n3. If a students current position is in between of a page,\n his/hers takhteets first day target will be\n pages per day assigned + the remaining current page.")
print("\n4. All the takhteets of students will be generated in the excel file\n 'test2.xlsx' including all details in arabic, while the program is being\n executed.\n Make sure that this excel file is closed while the program is being executed,\n and save both, the program and the file on desktop to avoid unwanted errors.")
print("\n5. User can either select the tabular layout or grid layout of takhteet.")
wait = input("\nPRESS ENTER TO CONTINUE.")

print ("\n" * 40)


w=0
studs=int(input("Enter number of students "))
year=int(input("Enter hijri year "))
yr=ayat(year)
for k in range(1,13):
    print(month[w]+" ."+str(k))
    w+=1
mon=int(input("Enter month no. with the help of above list "))
mont=month[mon-1]
tno=int(input("Enter total no.of working days of this month "))
form=int(input("Enter 1 for tabular or 2 for grid display of takhteet form in excel "))
print("\n")
z=1
while(studs>0):
    j=0
    print("\n")
    print("For student ",z,":")
    
    name=input("Enter student name ")
    n=float(input("Enter number of pages per day "))
    par=input("Is no.of pages progression of 0.5 y/n ")
    if(par=='y'):
        n=int(n*2)
        no=tno//2
    elif(par=='n'):
        int(n)
        no=tno
    for i in range(1,115):
        print(sup[j]+" ."+str(i))
        j+=1
    sur=int(input("Enter surat no. with the help of above list "))
    if(sur==82):
        sur=s.index("المطففين")+1
    elif(sur==86):
        sur=s.index("الأعلى")+1
    elif(sur==91):
        sur=s.index("البلد")+1
    elif(sur==93):
        sur=s.index("الشرح")+1
    elif(sur==95):
        sur=s.index("العلق")+1
    elif(sur==97):
        sur=s.index("البينة")+1
    elif(sur==99):
        sur=s.index("العاديات")+1
    elif(sur==101):
        sur=s.index("التكاثر")+1
    elif(sur==103 or sur==104):
        sur=s.index("الفيل")+1
    elif(sur==106 or sur==107):
        sur=s.index("الكوثر")+1
    elif(sur==109 or sur==110):
        sur=s.index("المسد")+1
    elif(sur==112 or sur==113):
        sur=s.index("الناس")+1
    sur-=1
    x=int(input("Enter current ayat number "))
    f=search(a[sur],x)
    f=int(f)
    if(f==53):
        sur=sur+1
        f=0
    print("**********************TAKHTEET*************************")
    print("Student name: ",name)
    print("Pages per day: ",ayat(n))
    print("\n\nالآية\t\tالسورة\n")
    generate(form,sur,f,no,n,name,par,tno,mont,yr)
    print("*******************************************************")
    studs-=1
    z=z+1
print(" معهد الزهراء-الشارقة")
print("\n")
print(" 'السهولة في التخطيط' ")
print("\n")
print("Developed by:-\n\t Mohammed Hasan Kothaliya")
time.sleep(5)



    

